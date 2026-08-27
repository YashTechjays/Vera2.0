"""Pure XLSX mapping layer for the form export — DB-free, format-agnostic inputs.

The workbook IS PHI (it carries field values): callers stream it inside an
authed, audited, no-store response and never log its contents. A future PDF
renderer consumes the same arguments.
"""

from collections.abc import Mapping, Sequence
from io import BytesIO
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from vera_core.forms.conditions import is_v2
from vera_core.forms.dsl import FormSchemaDoc
from vera_core.forms.export_form_sheet import autofit_columns, cell_str, render_form_sheet
from vera_core.services.call_provenance import CallAttempt, FieldProvenance

_BOLD = Font(bold=True)


def _bold_header(ws: Worksheet, title: str) -> None:
    """Append a blank separator row then a bold-title row to *ws*."""
    ws.append([])
    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = _BOLD


def _column_headings(ws: Worksheet, headings: list[str]) -> None:
    """Append a heading row with every cell bold — not just the first."""
    ws.append(headings)
    for col in range(1, len(headings) + 1):
        ws.cell(row=ws.max_row, column=col).font = _BOLD


def build_workbook(
    schema_json: Mapping[str, Any],
    values: Mapping[str, Any],
    sources: Mapping[str, str],
    provenance: Mapping[str, FieldProvenance],
    attempts: Sequence[CallAttempt],
) -> bytes:
    wb = Workbook()
    form_ws = cast(Worksheet, wb.active)
    form_ws.title = "Form"

    sorted_paths = sorted(values)

    if is_v2(schema_json):
        doc = FormSchemaDoc.model_validate(schema_json)
        # The renderer's single schema walk also yields the {path: title} map
        # that feeds the Provenance sheet — same content as before.
        titles: dict[str, str] = render_form_sheet(form_ws, doc, values)
    else:
        titles = {p: p for p in sorted_paths}
        for path in sorted_paths:
            form_ws.append([path, cell_str(values[path])])
        autofit_columns(form_ws, max_width=64.0)

    prov_ws = cast(Worksheet, wb.create_sheet("Provenance"))
    _column_headings(
        prov_ws,
        [
            "Field path",
            "Label",
            "Source",
            "Attempt",
            "Mode",
            "Judge confidence",
            "Supported",
            # False = the call that produced this value captured no rep reference number,
            # so nothing ties it to a payer-side record. The value is still current.
            "Authoritative",
        ],
    )
    prov_ws.freeze_panes = "A2"
    for path in sorted_paths:
        p = provenance.get(path)
        prov_ws.append(
            [
                path,
                titles.get(path, path),
                sources.get(path, ""),
                p.attempt if p else None,
                p.mode if p else None,
                p.judge.confidence if p and p.judge else None,
                p.judge.supported if p and p.judge else None,
                p.authoritative if p else None,
            ]
        )

    attempt_by_id = {a.id: a.attempt for a in attempts}
    _bold_header(prov_ws, "Call history")
    _column_headings(
        prov_ws, ["Attempt", "Mode", "Status", "Created at", "Retry of attempt", "Authoritative"]
    )
    for a in attempts:
        prov_ws.append(
            [
                a.attempt,
                a.mode,
                a.status,
                a.created_at.isoformat(),
                attempt_by_id.get(a.retry_of) if a.retry_of else None,
                a.authoritative,
            ]
        )

    # Field paths run long; the Form sheet sets its own bounds in render_form_sheet.
    autofit_columns(prov_ws, max_width=64.0)

    _neutralize_formulas(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _neutralize_formulas(wb: Workbook) -> None:
    """Force every cell to be data, never a live formula. openpyxl types any
    string starting with "=" as a formula (data_type "f"), and form values come
    from LLM extraction of rep speech, the intake API, and human edits — all
    attacker-influenceable. A value like '=HYPERLINK(...)' must open in Excel
    as inert text (PHI exfiltration surface), so retype it as a string; the
    stored text is unchanged."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    cell.data_type = "s"
