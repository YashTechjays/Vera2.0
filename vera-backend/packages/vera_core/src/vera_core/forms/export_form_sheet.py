"""UI-replica "Form" sheet for the XLSX export — a Python port of the
frontend's schema-derived layout rules.

The sheet carries PHI (field values): callers stream it inside an authed,
audited, no-store response. This module must never log.

Layout sources of truth (update together):
- placement / flow: vera-frontend/src/components/ibv/SchemaForm.tsx
- matrix model:     vera-frontend/src/lib/ibv/schema.ts (getSectionTable)
- grid headers:     vera-frontend/src/components/ibv/SectionMatrix.tsx
"""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from typing import Any

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from vera_core.forms.conditions import is_applicable, is_required, leaf_gates
from vera_core.forms.dsl import FormSchemaDoc, Group, Leaf, Section


@dataclass(frozen=True)
class TableCell:
    path: str
    leaf: Leaf


@dataclass(frozen=True)
class TableRow:
    path: str
    label: str
    cells: dict[str, TableCell]


@dataclass(frozen=True)
class TableGroup:
    path: str
    label: str
    icd10: str
    rows: list[TableRow]
    extras: dict[str, TableCell]


@dataclass(frozen=True)
class SectionTable:
    columns: list[tuple[str, str]]  # (leaf key, column title)
    extra_columns: list[tuple[str, str]]  # (leaf key, column title)
    has_icd: bool
    groups: list[TableGroup]
    leaves: list[tuple[str, Leaf]]  # section-level (path, leaf) field rows


def _leaf_entries(g: Group | Section) -> list[tuple[str, Leaf]]:
    return [(k, f) for k, f in g.fields.items() if isinstance(f, Leaf)]


def _group_entries(g: Group | Section) -> list[tuple[str, Group]]:
    return [(k, f) for k, f in g.fields.items() if isinstance(f, Group)]


def section_table(section_key: str, section: Section) -> SectionTable | None:
    """Matrix model for a ``ui.layout: "table"`` section — the layout hint alone
    decides. Top-level groups are bands; their subgroups are rows and their own
    leaves are per-group rowspan extras; a subgroup-less group is itself one row
    labelled by its CPT codes, its leaves split row-cells/extras by key."""
    if section.ui is None or section.ui.layout != "table":
        return None
    base = f"sections.{section_key}"
    top_leaves = [(f"{base}.{k}", f) for k, f in _leaf_entries(section)]
    top_groups = _group_entries(section)

    extra_titles: dict[str, str] = {}
    for _, g in top_groups:
        if not _group_entries(g):
            continue
        for k, f in _leaf_entries(g):
            extra_titles.setdefault(k, f.title)

    column_titles: dict[str, str] = {}
    groups: list[TableGroup] = []
    for gkey, g in top_groups:
        gpath = f"{base}.{gkey}"
        subgroups = _group_entries(g)
        extras: dict[str, TableCell] = {}
        rows: list[TableRow] = []
        if subgroups:
            for k, f in _leaf_entries(g):
                extras[k] = TableCell(f"{gpath}.{k}", f)
            for rkey, r in subgroups:
                rpath = f"{gpath}.{rkey}"
                entries = _leaf_entries(r)
                for k, f in entries:
                    column_titles.setdefault(k, f.title)
                rows.append(
                    TableRow(
                        rpath,
                        r.title,
                        {k: TableCell(f"{rpath}.{k}", f) for k, f in entries},
                    )
                )
        else:
            entries = _leaf_entries(g)
            row_entries = [(k, f) for k, f in entries if k not in extra_titles]
            for k, f in row_entries:
                column_titles.setdefault(k, f.title)
            for k, f in entries:
                if k in extra_titles:
                    extras[k] = TableCell(f"{gpath}.{k}", f)
            cpt = ", ".join(g.codes.cpt) if g.codes and g.codes.cpt is not None else "—"
            rows.append(
                TableRow(
                    gpath,
                    cpt,
                    {k: TableCell(f"{gpath}.{k}", f) for k, f in row_entries},
                )
            )
        icd = ", ".join(g.codes.icd10) if g.codes and g.codes.icd10 else ""
        groups.append(TableGroup(gpath, g.title, icd, rows, extras))

    return SectionTable(
        columns=list(column_titles.items()),
        extra_columns=list(extra_titles.items()),
        has_icd=any(g.icd10 for g in groups),
        groups=groups,
        leaves=top_leaves,
    )


# Placement mirrors vera-frontend/src/components/ibv/SchemaForm.tsx
# (LEFT_TOP / RIGHT_TOP / RAIL) — update both together.
LEFT_TOP = ["patient_information", "insurance_information"]
RIGHT_TOP = ["appointment_information", "verification_information", "benefit_coverage"]
RAIL = [
    "hospital_information",
    "provider_reference_information",
    "insurance_reference_information",
]

# Column anchors (1-based): left block A-B, right D-E, rail G-H; spacers C, F.
_LEFT_COL, _RIGHT_COL, _RAIL_COL = 1, 4, 7

_THIN = Side(style="thin", color="B0B0B0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_TOP_LEFT = Alignment(vertical="top", wrap_text=True)
_FILL_CONTEXT = PatternFill("solid", start_color="C6EFCE")  # UI's green header
_FILL_PLAIN = PatternFill("solid", start_color="E7E6E6")
_FILL_GRID_HEADER = PatternFill("solid", start_color="F2F2F2")
_FILL_INAPPLICABLE = PatternFill("solid", start_color="F5F5F5")


def _str(v: Any) -> str:
    """Coerce a field value to a spreadsheet-safe string; None → empty string."""
    return "" if v is None else str(v)


class _Ctx:
    """Render context: values + applicability machinery, computed once."""

    def __init__(self, doc: FormSchemaDoc, values: Mapping[str, Any]) -> None:
        self.values = values
        self.shared = doc.shared_conditions or {}
        self.gates = {path: gates for path, _leaf, gates in leaf_gates(doc)}

    def applicable(self, path: str) -> bool:
        return is_applicable(self.gates.get(path, ()), self.values, self.shared)


def _style(
    ws: Worksheet,
    row: int,
    col: int,
    *,
    fill: PatternFill | None = None,
    bold: bool = False,
    center: bool = False,
) -> None:
    c = ws.cell(row=row, column=col)
    c.border = _BORDER
    if fill is not None:
        c.fill = fill
    if bold:
        c.font = _BOLD
    c.alignment = _CENTER if center else _TOP_LEFT


def _vmerge(ws: Worksheet, top: int, col: int, n: int) -> None:
    """Merge a single column down *n* rows starting at *top* (no-op for n <= 1)."""
    if n > 1:
        ws.merge_cells(start_row=top, start_column=col, end_row=top + n - 1, end_column=col)


def _title_bar(ws: Worksheet, row: int, col: int, span: int, text: str, *, context: bool) -> None:
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    ws.cell(row=row, column=col, value=text)
    fill = _FILL_CONTEXT if context else _FILL_PLAIN
    for i in range(span):
        _style(ws, row, col + i, fill=fill, bold=True, center=True)


def _leaf_row(ws: Worksheet, row: int, col: int, path: str, leaf: Leaf, ctx: _Ctx) -> None:
    applicable = ctx.applicable(path)
    star = " *" if applicable and is_required(leaf, ctx.values, ctx.shared) else ""
    ws.cell(row=row, column=col, value=f"{leaf.title}{star}")
    ws.cell(row=row, column=col + 1, value=_str(ctx.values.get(path)) if applicable else "")
    fill = None if applicable else _FILL_INAPPLICABLE
    _style(ws, row, col, fill=fill, bold=True)
    _style(ws, row, col + 1, fill=fill)


def _field_block(
    ws: Worksheet, row: int, col: int, section_key: str, section: Section, ctx: _Ctx
) -> int:
    """Label/value section anchored at (row, col); returns rows consumed.
    Mirrors the FE flattenSection order: groups emit a sub-header then their
    children, depth-first."""
    _title_bar(ws, row, col, 2, section.title, context=section.role == "context")
    r = row + 1

    def emit(prefix: str, fields: Mapping[str, Any]) -> None:
        nonlocal r
        for key, f in fields.items():
            path = f"{prefix}.{key}"
            if isinstance(f, Leaf):
                _leaf_row(ws, r, col, path, f, ctx)
                r += 1
            else:  # nested group: sub-header row, then children
                ws.cell(row=r, column=col, value=f.title)
                ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + 1)
                _style(ws, r, col, bold=True)
                _style(ws, r, col + 1)
                r += 1
                emit(path, f.fields)

    emit(f"sections.{section_key}", section.fields)
    return r - row


def _grid_block(ws: Worksheet, row: int, section_key: str, section: Section, ctx: _Ctx) -> int:
    """Full-width matrix section; returns rows consumed. Header order mirrors
    SectionMatrix.tsx: Service, [ICD-10], CPT Code, columns…, extras…"""
    table = section_table(section_key, section)
    assert table is not None  # caller checked ui.layout
    static = ["Service"] + (["ICD-10"] if table.has_icd else []) + ["CPT Code"]
    width = len(static) + len(table.columns) + len(table.extra_columns)

    _title_bar(ws, row, 1, width, section.title, context=section.role == "context")
    r = row + 1
    for path, leaf in table.leaves:  # section-level field rows above the grid
        _leaf_row(ws, r, 1, path, leaf, ctx)
        r += 1

    headers = static + [t for _, t in table.columns] + [t for _, t in table.extra_columns]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=1 + i, value=h)
        _style(ws, r, 1 + i, fill=_FILL_GRID_HEADER, bold=True, center=True)
    r += 1

    icd_col = 2 if table.has_icd else None
    cpt_col = 3 if table.has_icd else 2
    first_val_col = cpt_col + 1
    extra_col0 = first_val_col + len(table.columns)

    band_cols = [
        1,
        *((icd_col,) if icd_col is not None else ()),
        cpt_col,
        *range(extra_col0, extra_col0 + len(table.extra_columns)),
    ]
    for group in table.groups:
        n = len(group.rows)
        top = r
        ws.cell(row=top, column=1, value=group.label)
        _vmerge(ws, top, 1, n)
        if icd_col is not None:
            ws.cell(row=top, column=icd_col, value=group.icd10)
            _vmerge(ws, top, icd_col, n)
        for j, (key, _title) in enumerate(table.extra_columns):
            cell = group.extras.get(key)
            col = extra_col0 + j
            if cell is not None and ctx.applicable(cell.path):
                ws.cell(row=top, column=col, value=_str(ctx.values.get(cell.path)))
            _vmerge(ws, top, col, n)
        for row_model in group.rows:
            ws.cell(row=r, column=cpt_col, value=row_model.label)
            for j, (key, _title) in enumerate(table.columns):
                cell = row_model.cells.get(key)
                col = first_val_col + j
                applicable = cell is not None and ctx.applicable(cell.path)
                if cell is not None and applicable:
                    ws.cell(row=r, column=col, value=_str(ctx.values.get(cell.path)))
                _style(ws, r, col, fill=None if applicable else _FILL_INAPPLICABLE)
            r += 1
        # Border/style pass over the band's static + extra cells.
        for rr in range(top, top + n):
            for col in band_cols:
                _style(ws, rr, col, bold=col == 1)
    return r - row


def render_form_sheet(ws: Worksheet, doc: FormSchemaDoc, values: Mapping[str, Any]) -> None:
    """Write the UI-replica form into *ws* (title, widths, sections)."""
    ws.title = "Form"
    ctx = _Ctx(doc, values)
    sections = doc.sections

    def stack(keys: list[str], col: int) -> int:
        r = 1
        for key in keys:
            if key not in sections:
                continue
            r += _field_block(ws, r, col, key, sections[key], ctx) + 1
        return r

    next_row = max(
        stack(LEFT_TOP, _LEFT_COL),
        stack(RIGHT_TOP, _RIGHT_COL),
        stack(RAIL, _RAIL_COL),
    )

    placed = set(LEFT_TOP) | set(RIGHT_TOP) | set(RAIL)
    rest = [k for k in sections if k not in placed]
    r = next_row + 1
    run: list[str] = []

    def flush_run() -> None:
        nonlocal r, run
        for i in range(0, len(run), 2):
            pair = run[i : i + 2]
            heights = []
            for j, key in enumerate(pair):
                col = _LEFT_COL if j == 0 else _RIGHT_COL
                heights.append(_field_block(ws, r, col, key, sections[key], ctx))
            r += max(heights) + 1
        run = []

    for key in rest:
        sec = sections[key]
        if sec.ui is not None and sec.ui.layout == "table":
            flush_run()
            r += _grid_block(ws, r, key, sec, ctx) + 1
        else:
            run.append(key)
    flush_run()

    widths = {1: 30, 2: 22, 3: 14, 4: 30, 5: 22, 6: 3, 7: 30, 8: 22}
    for idx, w in widths.items():
        ws.column_dimensions[chr(ord("A") + idx - 1)].width = w
