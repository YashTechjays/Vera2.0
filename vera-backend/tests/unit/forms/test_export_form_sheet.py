"""Unit tests for the UI-replica export sheet — pure, no DB."""

from __future__ import annotations

import copy
from pathlib import Path
from typing import Any, cast

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from vera_core.forms.catalog import SCHEMAS
from vera_core.forms.conditions import alternative_fills, leaf_gates
from vera_core.forms.dsl import FormSchemaDoc, load_document
from vera_core.forms.export_form_sheet import (
    LEFT_TOP,
    RAIL,
    RIGHT_TOP,
    _Ctx,
    _leaf_value,
    render_form_sheet,
    section_table,
)

# Same depth as test_schema_dsl.py's FORM_SCHEMA_DIR: tests/unit/forms/<file> -> vera-backend.
FORM_SCHEMA_DIR = Path(__file__).resolve().parents[3] / "data" / "form_schemas"

# A minimal v2 doc with one table section exercising BOTH group shapes:
#  - ivf: subgroup rows (cpt_1, cpt_2) + group-level extras (cycle_limit, notes)
#  - oi:  leaf-only group (its leaves split into row cells vs extras by key)
_TABLE_DOC: dict[str, Any] = {
    "dsl_version": "2.1",
    "name": "Replica Test Schema",
    "insurance_type": "infertility_treatment",
    "system_fields": {"in_network": "sections.info.in_network"},
    "rep_call_reference_number_field": "sections.info.in_network",
    "promoted_fields": {},  # filled in fixture below
    "sections": {
        "info": {
            "title": "Info",
            "role": "collect",
            "fields": {
                "in_network": {
                    "type": "text",
                    "title": "In Network",
                    "role": "ask",
                    "required": True,
                    "prompt": {"ask": "In network?"},
                },
            },
        },
        "treatment": {
            "title": "Treatment",
            "role": "collect",
            "ui": {"layout": "table"},
            "fields": {
                "tx_covered": {
                    "type": "text",
                    "title": "Treatment Covered",
                    "role": "ask",
                    "prompt": {"ask": "Covered?"},
                },
                "ivf": {
                    "type": "group",
                    "title": "In Vitro Fertilization (IVF)",
                    "codes": {"icd10": ["Z31.83"]},
                    "fields": {
                        "cycle_limit": {
                            "type": "text",
                            "title": "Cycle Limit",
                            "role": "ask",
                            "prompt": {"ask": "Cycle limit?"},
                        },
                        "notes": {
                            "type": "text",
                            "title": "Additional Notes",
                            "role": "ask",
                            "prompt": {"ask": "Notes?"},
                        },
                        "cpt_58970": {
                            "type": "group",
                            "title": "CPT 58970",
                            "fields": {
                                "covered": {
                                    "type": "text",
                                    "title": "Covered",
                                    "role": "ask",
                                    "prompt": {"ask": "Covered?"},
                                },
                                "copay": {
                                    "type": "text",
                                    "title": "Copay ($)",
                                    "role": "ask",
                                    "prompt": {"ask": "Copay?"},
                                },
                            },
                        },
                        "cpt_89280": {
                            "type": "group",
                            "title": "CPT 89280",
                            "fields": {
                                "covered": {
                                    "type": "text",
                                    "title": "Covered",
                                    "role": "ask",
                                    "prompt": {"ask": "Covered?"},
                                },
                                "copay": {
                                    "type": "text",
                                    "title": "Copay ($)",
                                    "role": "ask",
                                    "prompt": {"ask": "Copay?"},
                                },
                            },
                        },
                    },
                },
                "oi": {
                    "type": "group",
                    "title": "Ovulation Induction",
                    "codes": {"icd10": ["N97.0"], "cpt": ["58323"]},
                    "fields": {
                        "covered": {
                            "type": "text",
                            "title": "Covered",
                            "role": "ask",
                            "prompt": {"ask": "Covered?"},
                        },
                        "copay": {
                            "type": "text",
                            "title": "Copay ($)",
                            "role": "ask",
                            "prompt": {"ask": "Copay?"},
                        },
                        "cycle_limit": {
                            "type": "text",
                            "title": "Cycle Limit",
                            "role": "ask",
                            "prompt": {"ask": "Cycle limit?"},
                        },
                    },
                },
            },
        },
    },
    "tasks": [{"task_key": "main", "title": "Main", "sections": ["info", "treatment"]}],
}


def _doc() -> FormSchemaDoc:
    from vera_core.forms.dsl import PromotedFields

    raw = dict(_TABLE_DOC)
    raw["promoted_fields"] = dict.fromkeys(PromotedFields.model_fields, "sections.info.in_network")
    return FormSchemaDoc.model_validate(raw)


def test_section_table_none_for_non_table_sections() -> None:
    doc = _doc()
    assert section_table("info", doc.sections["info"]) is None


def test_section_table_model_matches_frontend_rules() -> None:
    doc = _doc()
    table = section_table("treatment", doc.sections["treatment"])
    assert table is not None

    # Section-level leaves render above the grid.
    assert [p for p, _ in table.leaves] == ["sections.treatment.tx_covered"]

    # Columns = shared leaf keys of subgroup rows, first-seen order.
    assert [key for key, _ in table.columns] == ["covered", "copay"]
    assert [t for _, t in table.columns] == ["Covered", "Copay ($)"]
    # Extras = leaf keys sitting beside subgroups inside any group.
    assert [key for key, _ in table.extra_columns] == ["cycle_limit", "notes"]
    assert table.has_icd is True

    ivf, oi = table.groups
    assert ivf.label == "In Vitro Fertilization (IVF)"
    assert ivf.icd10 == "Z31.83"
    assert [r.label for r in ivf.rows] == ["CPT 58970", "CPT 89280"]
    assert ivf.rows[0].cells["copay"].path == ("sections.treatment.ivf.cpt_58970.copay")
    assert set(ivf.extras) == {"cycle_limit", "notes"}

    # Leaf-only group: the group itself is one row labelled by its CPT codes;
    # keys that exist as extras elsewhere split out of the row cells.
    assert oi.label == "Ovulation Induction"
    assert len(oi.rows) == 1
    assert oi.rows[0].label == "58323"
    assert set(oi.rows[0].cells) == {"covered", "copay"}
    assert set(oi.extras) == {"cycle_limit"}


# Extends the table fixture with placed sections: two LEFT_TOP, one RIGHT_TOP
# (context role → green), one RAIL (context), plus a trailing two-up pair.
#
# NOTE: two shapes were adapted from a first draft of this fixture:
#  - `applicable_when` uses the DSL's actual `Comparison` field name (`field`),
#    not `path` — see dsl.py's `Comparison` model.
#  - `tasks[].sections` lists only `collect`-role sections (patient_information,
#    treatment, alpha, beta); the DSL validator rejects a `context` section
#    (appointment_information, hospital_information) inside a task's section
#    list ("only collect sections belong to tasks"). The gating INTENT — spouse
#    row gated on patient_name == "married" — is unchanged.
_PLACED_DOC: dict[str, Any] = {
    **_TABLE_DOC,
    "sections": {
        "patient_information": {
            "title": "Patient Information",
            "role": "collect",
            "fields": {
                "patient_name": {
                    "type": "text",
                    "title": "Patient Name",
                    "role": "ask",
                    "required": True,
                    "prompt": {"ask": "Name?"},
                },
                "spouse_name": {
                    "type": "text",
                    "title": "Spouse Name",
                    "role": "ask",
                    "applicable_when": {
                        "field": "sections.patient_information.patient_name",
                        "op": "eq",
                        "value": "married",
                    },
                    "prompt": {"ask": "Spouse?"},
                },
            },
        },
        "appointment_information": {
            "title": "Appointment Information",
            "role": "context",
            "fields": {
                "appointment_date": {
                    "type": "text",
                    "title": "Appointment Date",
                    "role": "context",
                },
            },
        },
        "hospital_information": {
            "title": "Hospital Information",
            "role": "context",
            "fields": {
                "hospital_name": {
                    "type": "text",
                    "title": "Hospital Name",
                    "role": "context",
                },
            },
        },
        "treatment": _TABLE_DOC["sections"]["treatment"],
        "alpha": {
            "title": "Alpha",
            "role": "collect",
            "fields": {
                "a1": {"type": "text", "title": "A One", "role": "ask", "prompt": {"ask": "?"}},
            },
        },
        "beta": {
            "title": "Beta",
            "role": "collect",
            "fields": {
                "b1": {"type": "text", "title": "B One", "role": "ask", "prompt": {"ask": "?"}},
            },
        },
    },
    "system_fields": {"in_network": "sections.patient_information.patient_name"},
    "rep_call_reference_number_field": "sections.patient_information.patient_name",
    "tasks": [
        {
            "task_key": "main",
            "title": "Main",
            "sections": ["patient_information", "treatment", "alpha", "beta"],
        }
    ],
}


def _placed_doc() -> FormSchemaDoc:
    from vera_core.forms.dsl import PromotedFields

    raw = dict(_PLACED_DOC)
    raw["promoted_fields"] = dict.fromkeys(
        PromotedFields.model_fields, "sections.patient_information.patient_name"
    )
    return FormSchemaDoc.model_validate(raw)


def _render(values: dict[str, Any]) -> Worksheet:
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    render_form_sheet(ws, _placed_doc(), values)
    return ws


def test_top_band_geometry_and_values() -> None:
    ws = _render({"sections.patient_information.patient_name": "Jane"})
    # Left block anchored at A1; right at D1; rail at G1.
    assert ws.cell(row=1, column=1).value == "Patient Information"
    assert ws.cell(row=1, column=4).value == "Appointment Information"
    assert ws.cell(row=1, column=7).value == "Hospital Information"
    # Label + value adjacency, with the required marker.
    assert ws.cell(row=2, column=1).value == "Patient Name *"
    assert ws.cell(row=2, column=2).value == "Jane"


def test_context_sections_get_green_title_fill() -> None:
    ws = _render({})
    ctx = ws.cell(row=1, column=4).fill.start_color.rgb  # context section
    plain = ws.cell(row=1, column=1).fill.start_color.rgb  # collect section
    assert ctx != plain
    assert str(ctx).endswith("22C55E")


def test_inapplicable_leaf_grayed_with_empty_value() -> None:
    ws = _render({"sections.patient_information.spouse_name": "should-not-show"})
    # patient_name != "married" → spouse row (row 3, left block) is gated off.
    assert ws.cell(row=3, column=2).value in (None, "")
    assert str(ws.cell(row=3, column=2).fill.start_color.rgb).endswith("F7F7F7")


def test_grid_full_width_and_two_up_flow_below_band() -> None:
    ws = _render({"sections.treatment.ivf.cpt_58970.copay": "30"})
    # Find the grid title and header row.
    titles = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    grid_row = titles["Treatment"]
    header = grid_row + 2  # +1 section-leaf row (tx_covered) sits between
    assert ws.cell(row=header, column=1).value == "Service"
    assert ws.cell(row=header, column=2).value == "ICD-10"
    assert ws.cell(row=header, column=3).value == "CPT Code"
    assert ws.cell(row=header, column=4).value == "Covered"
    assert ws.cell(row=header, column=5).value == "Copay ($)"
    assert ws.cell(row=header, column=6).value == "Cycle Limit"
    assert ws.cell(row=header, column=7).value == "Additional Notes"
    # IVF band: Service cell merged across its two CPT rows; copay value lands.
    ivf_first = header + 1
    merged = {str(rng) for rng in ws.merged_cells.ranges}
    assert f"A{ivf_first}:A{ivf_first + 1}" in merged
    assert ws.cell(row=ivf_first, column=1).value == "In Vitro Fertilization (IVF)"
    assert ws.cell(row=ivf_first, column=3).value == "CPT 58970"
    assert ws.cell(row=ivf_first, column=5).value == "30"
    # Two-up run: alpha (left band) and beta (right band) share a row.
    alpha_row = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Alpha"
    )
    assert ws.cell(row=alpha_row, column=4).value == "Beta"


def test_inapplicable_grid_extras_cell_grayed_with_empty_value() -> None:
    from vera_core.forms.dsl import PromotedFields

    # Gate the ivf group's cycle_limit extras leaf on a value that's absent, so
    # its rowspan extras cell is gated off; oi's cycle_limit stays applicable.
    raw = copy.deepcopy(_PLACED_DOC)
    cycle_limit = raw["sections"]["treatment"]["fields"]["ivf"]["fields"]["cycle_limit"]
    cycle_limit["applicable_when"] = {
        "field": "sections.patient_information.patient_name",
        "op": "eq",
        "value": "gate-open",
    }
    raw["promoted_fields"] = dict.fromkeys(
        PromotedFields.model_fields, "sections.patient_information.patient_name"
    )
    doc = FormSchemaDoc.model_validate(raw)
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    render_form_sheet(
        ws,
        doc,
        {
            "sections.treatment.ivf.cycle_limit": "should-not-show",
            "sections.treatment.oi.cycle_limit": "4",
        },
    )
    titles = {ws.cell(row=r, column=1).value: r for r in range(1, ws.max_row + 1)}
    header = titles["Treatment"] + 2  # +1 section-leaf row (tx_covered)
    assert ws.cell(row=header, column=6).value == "Cycle Limit"
    ivf_first = header + 1  # ivf band: rows ivf_first..ivf_first+1
    oi_row = ivf_first + 2  # oi band: single row below the ivf band
    # Gated ivf extras cell: empty AND grayed, like every other gated leaf.
    assert ws.cell(row=ivf_first, column=6).value in (None, "")
    assert str(ws.cell(row=ivf_first, column=6).fill.start_color.rgb).endswith("F7F7F7")
    # Applicable oi extras cell: value lands, no gray fill.
    assert ws.cell(row=oi_row, column=6).value == "4"
    assert str(ws.cell(row=oi_row, column=6).fill.start_color.rgb).endswith("D0E0E3")


def test_ibv_standard_renders_and_placement_lists_exist() -> None:
    """Drift/smoke guard against the real, shipped ibv_standard schema: the
    placement lists (LEFT_TOP/RIGHT_TOP/RAIL) must still reference sections
    that exist in the committed artifact, and rendering it end-to-end must
    not raise."""
    filename, _build = SCHEMAS["infertility_treatment"]
    text = (FORM_SCHEMA_DIR / filename).read_text()
    doc = load_document(text)
    for key in (*LEFT_TOP, *RIGHT_TOP, *RAIL):
        assert key in doc.sections, f"placement list references missing section {key}"
    wb = Workbook()
    render_form_sheet(cast(Worksheet, wb.active), doc, {})


def test_placement_constants_reference_fe() -> None:
    # Guard: the constants stay aligned with SchemaForm.tsx's lists.
    assert LEFT_TOP == ["patient_information", "insurance_information"]
    assert RIGHT_TOP == [
        "appointment_information",
        "verification_information",
        "benefit_coverage",
    ]
    assert RAIL == [
        "hospital_information",
        "provider_reference_information",
        "insurance_reference_information",
    ]


def test_every_catalog_schema_renders() -> None:
    """Smoke over ALL shipped schemas (not just ibv_standard) — a schema
    without the top-band sections must still render via the below-band flow."""
    for filename, _build in SCHEMAS.values():
        doc = load_document((FORM_SCHEMA_DIR / filename).read_text())
        wb = Workbook()
        render_form_sheet(cast(Worksheet, wb.active), doc, {})


def test_declared_default_counts_as_filled_on_export() -> None:
    """DSL §4.4: a leaf with a declared default and no stored value exports the
    default (field blocks AND grid cells), matching completion_pct_v2."""
    raw = copy.deepcopy(_PLACED_DOC)
    fields = raw["sections"]["patient_information"]["fields"]
    fields["patient_name"]["default"] = "N/A"
    grid = raw["sections"]["treatment"]["fields"]
    grid["ivf"]["fields"]["cpt_58970"]["fields"]["copay"]["default"] = "0"
    from vera_core.forms.dsl import PromotedFields

    raw["promoted_fields"] = dict.fromkeys(
        PromotedFields.model_fields, "sections.patient_information.patient_name"
    )
    doc = FormSchemaDoc.model_validate(raw)
    wb = Workbook()
    ws = cast(Worksheet, wb.active)
    render_form_sheet(ws, doc, {})

    assert ws.cell(row=2, column=2).value == "N/A"  # field block default
    grid_title = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "Treatment"
    )
    header = grid_title + 2
    assert ws.cell(row=header + 1, column=5).value == "0"  # grid cell default


def test_ui_palette_usage_tints_and_value_fill() -> None:
    """Label cells carry the UI's usage tints (usageMeta.ts) and value cells
    the teal input background (--color-ibv-input-bg)."""
    ws = _render({"sections.patient_information.patient_name": "Jane"})
    # patient_name is a system_fields target -> violet-100 label tint.
    assert str(ws.cell(row=2, column=1).fill.start_color.rgb).endswith("EDE9FE")
    # its value cell uses the teal input background.
    assert str(ws.cell(row=2, column=2).fill.start_color.rgb).endswith("D0E0E3")
    # appointment_date is role=context -> green-100 label tint (right band).
    assert str(ws.cell(row=2, column=4).fill.start_color.rgb).endswith("DCFCE7")


_IUI = "sections.infertility_treatment.intrauterine_insemination.cpt_58323"


def test_a_filled_either_or_side_reaches_the_sheet_and_is_not_greyed() -> None:
    """The whole point of writing the fill rather than showing a placeholder: the export is the
    platform's final product, and an inapplicable cell is written blank and grey."""
    doc = load_document((FORM_SCHEMA_DIR / "ibv_form_standard_v2.json").read_text("utf-8"))
    answered = {
        "sections.infertility_treatment.infertility_tx_covered": "Yes",
        f"{_IUI}.covered": "Yes",
        f"{_IUI}.coinsurance": "30",
    }
    fills = alternative_fills(doc, answered, f"{_IUI}.coinsurance")
    assert fills == {f"{_IUI}.copay": "$0"}

    ctx = _Ctx(doc, {**answered, **fills})
    leaves = {path: leaf for path, leaf, _gates in leaf_gates(doc)}
    # Applicable, so the renderer writes the value instead of blanking and graying the cell.
    assert ctx.applicable(f"{_IUI}.copay")
    assert _leaf_value(f"{_IUI}.copay", leaves[f"{_IUI}.copay"], ctx) == "$0"
    # The rep's own answer is untouched on the other side of the pair.
    assert _leaf_value(f"{_IUI}.coinsurance", leaves[f"{_IUI}.coinsurance"], ctx) == "30"
