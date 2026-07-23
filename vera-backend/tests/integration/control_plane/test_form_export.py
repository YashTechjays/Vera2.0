"""Integration tests for POST /patient-forms/{form_id}/export.

Tests:
- 200 response with XLSX workbook + ledger row written
- 422 when form is not completed
- 403 when caller lacks forms:export permission
"""

from collections.abc import AsyncGenerator
from io import BytesIO
from uuid import UUID

import httpx
import pytest
from openpyxl import load_workbook
from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from tests.integration.control_plane.conftest import RBACWorld
from vera_core.db import uuid7
from vera_core.models.authoring import FormSchema, SchemaVersion
from vera_core.models.call import Call
from vera_core.models.enums import AnswerSource, CallStatus, FormStatus, InsuranceType
from vera_core.models.field_answer import CallFormSnapshot, FieldAnswer
from vera_core.models.patient_form import PatientForm

pytestmark = pytest.mark.integration


@pytest.fixture
async def completed_form_id(
    database_url: str,
    rbac_world: RBACWorld,
) -> AsyncGenerator[UUID]:
    """Seed a COMPLETED form under rbac_world.tenant_id.

    Seeded data:
    - SchemaVersion version=996
    - PatientForm (status=COMPLETED)
    - call1: full, completed, snapshot {} -> {'exp.a': 'hello'}
    - FieldAnswer(exp.a, ai_call, call1, is_current=True)

    Teardown in FK order; does NOT delete the tenant (owned by rbac_world).
    """
    tenant_id = rbac_world.tenant_id
    form_id = uuid7()
    call1_id = uuid7()
    schema_version_id = uuid7()

    engine = create_async_engine(database_url)
    sm: async_sessionmaker[AsyncSession] = async_sessionmaker(engine, expire_on_commit=False)

    schema_id: UUID
    schema_id_to_delete: UUID | None = None

    try:
        async with sm() as session, session.begin():
            # find-or-create FormSchema (UNIQUE on insurance_type)
            existing = (
                await session.execute(
                    select(FormSchema).where(
                        FormSchema.insurance_type == InsuranceType.DISEASE_ONLY.value
                    )
                )
            ).scalar_one_or_none()
            if existing is None:
                fs = FormSchema(
                    id=uuid7(),
                    insurance_type=InsuranceType.DISEASE_ONLY.value,
                    name="Export Test Schema",
                )
                session.add(fs)
                await session.flush()
                schema_id = fs.id
                schema_id_to_delete = fs.id
            else:
                schema_id = existing.id

            session.add(
                SchemaVersion(
                    id=schema_version_id,
                    schema_id=schema_id,
                    version=996,
                    schema_json={},
                )
            )
            session.add(
                PatientForm(
                    id=form_id,
                    tenant_id=tenant_id,
                    schema_version_id=schema_version_id,
                    patient_name="Export Patient",
                    status=FormStatus.COMPLETED.value,
                )
            )
            await session.flush()

            session.add(
                Call(
                    id=call1_id,
                    tenant_id=tenant_id,
                    form_id=form_id,
                    current_status=CallStatus.COMPLETED.value,
                    mode="full",
                )
            )
            await session.flush()

            session.add(
                CallFormSnapshot(
                    tenant_id=tenant_id,
                    call_id=call1_id,
                    before_state={},
                    after_state={"exp.a": "hello"},
                )
            )
            session.add(
                FieldAnswer(
                    id=uuid7(),
                    tenant_id=tenant_id,
                    form_id=form_id,
                    field_path="exp.a",
                    value={"value": "hello"},
                    source=AnswerSource.AI_CALL.value,
                    call_id=call1_id,
                    is_current=True,
                )
            )

        yield form_id

    finally:
        async with sm() as session, session.begin():
            # Teardown in FK order (export_artifact before patient_form)
            await session.execute(
                text("DELETE FROM export_artifact WHERE form_id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text(
                    "DELETE FROM field_evaluation WHERE answer_id IN "
                    "(SELECT id FROM field_answer WHERE form_id = :fid)"
                ).bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM field_answer WHERE form_id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text(
                    "DELETE FROM call_form_snapshot WHERE call_id IN "
                    "(SELECT id FROM call WHERE form_id = :fid)"
                ).bindparams(fid=form_id)
            )
            await session.execute(
                text(
                    "DELETE FROM call_lineage WHERE retry_call_id IN "
                    "(SELECT id FROM call WHERE form_id = :fid)"
                ).bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM call WHERE form_id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM patient_form WHERE id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM schema_version WHERE id = :sid").bindparams(sid=schema_version_id)
            )
            if schema_id_to_delete is not None:
                await session.execute(
                    text("DELETE FROM form_schema WHERE id = :fsid").bindparams(
                        fsid=schema_id_to_delete
                    )
                )
        await engine.dispose()


@pytest.fixture
async def non_completed_form_id(
    database_url: str,
    rbac_world: RBACWorld,
) -> AsyncGenerator[UUID]:
    """Seed an AI_PROCESSING (non-completed) form for the rejection test."""
    tenant_id = rbac_world.tenant_id
    form_id = uuid7()
    schema_version_id = uuid7()

    engine = create_async_engine(database_url)
    sm: async_sessionmaker[AsyncSession] = async_sessionmaker(engine, expire_on_commit=False)

    schema_id: UUID
    schema_id_to_delete: UUID | None = None

    try:
        async with sm() as session, session.begin():
            existing = (
                await session.execute(
                    select(FormSchema).where(
                        FormSchema.insurance_type == InsuranceType.DISEASE_ONLY.value
                    )
                )
            ).scalar_one_or_none()
            if existing is None:
                fs = FormSchema(
                    id=uuid7(),
                    insurance_type=InsuranceType.DISEASE_ONLY.value,
                    name="Non-Completed Export Test Schema",
                )
                session.add(fs)
                await session.flush()
                schema_id = fs.id
                schema_id_to_delete = fs.id
            else:
                schema_id = existing.id

            session.add(
                SchemaVersion(
                    id=schema_version_id,
                    schema_id=schema_id,
                    version=997,
                    schema_json={},
                )
            )
            session.add(
                PatientForm(
                    id=form_id,
                    tenant_id=tenant_id,
                    schema_version_id=schema_version_id,
                    patient_name="Non-Completed Patient",
                    status=FormStatus.AI_PROCESSING.value,
                )
            )

        yield form_id

    finally:
        async with sm() as session, session.begin():
            await session.execute(
                text("DELETE FROM export_artifact WHERE form_id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM patient_form WHERE id = :fid").bindparams(fid=form_id)
            )
            await session.execute(
                text("DELETE FROM schema_version WHERE id = :sid").bindparams(sid=schema_version_id)
            )
            if schema_id_to_delete is not None:
                await session.execute(
                    text("DELETE FROM form_schema WHERE id = :fsid").bindparams(
                        fsid=schema_id_to_delete
                    )
                )
        await engine.dispose()


@pytest.mark.asyncio
async def test_export_streams_xlsx_and_writes_ledger(
    client: httpx.AsyncClient,
    rbac_world: RBACWorld,
    completed_form_id: UUID,
    admin_sessionmaker: async_sessionmaker[AsyncSession],
) -> None:
    """completed_form_id: a COMPLETED form under rbac_world.tenant_id with at
    least one current answer and one call with snapshot."""
    resp = await client.post(
        f"/api/v1/patient-forms/{completed_form_id}/export",
        headers={"Authorization": f"Bearer {rbac_world.admin_token}"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert f"ibv-{completed_form_id}.xlsx" in resp.headers["content-disposition"]
    assert resp.headers["cache-control"] == "no-store"
    wb = load_workbook(BytesIO(resp.content))
    assert wb.sheetnames == ["Form", "Provenance"]

    async with admin_sessionmaker() as session:
        row = (
            await session.execute(
                text(
                    "SELECT format, sha256, gcs_uri FROM export_artifact WHERE form_id = :fid"
                ).bindparams(fid=completed_form_id)
            )
        ).one()
    assert row.format == "xlsx" and len(row.sha256) == 64 and row.gcs_uri is None


@pytest.mark.asyncio
async def test_export_rejects_non_completed(
    client: httpx.AsyncClient, rbac_world: RBACWorld, non_completed_form_id: UUID
) -> None:
    """Form that is not COMPLETED must be rejected with 422."""
    resp = await client.post(
        f"/api/v1/patient-forms/{non_completed_form_id}/export",
        headers={"Authorization": f"Bearer {rbac_world.admin_token}"},
    )
    # DefaultExceptionCode.VALIDATION_ERROR maps to 422
    assert resp.status_code == 422


@pytest.mark.asyncio
async def test_export_requires_permission(
    client: httpx.AsyncClient, rbac_world: RBACWorld, completed_form_id: UUID
) -> None:
    """A user without forms:export permission must receive 403."""
    resp = await client.post(
        f"/api/v1/patient-forms/{completed_form_id}/export",
        headers={"Authorization": f"Bearer {rbac_world.norole_token}"},
    )
    assert resp.status_code == 403
